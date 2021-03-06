#import stocklookup.py

import pandas as pd
from openpyxl.reader.excel import load_workbook

def routine(number):
   # data = pd.read_excel("NAICSData.xlsx")

    ##companies = data[data.NAICS1 == number]
    
    ##if()
    #print("Relevant companies".format(
    #len(companies)))

    #start here
    dictionary = {'comp0': None, 'comp1': None, 'comp2': None, 'comp3': None, 'comp4': None, 'comp5': None, 'comp6': None, 'comp7': None, 'comp8': None, 'comp9': None }
    book = load_workbook(filename = "NAICSData.xlsx")
    sheet = book['Sheet1']
    compNum = 0
    for row in range(2, 20):
        if compNum == 10:
            break
        if (sheet.cell(row, 7)).value == str(number):

            dictionary['comp' + str(compNum)] = sheet.cell(row, 3).value
            dictionary.update({'duns' + str(compNum) : sheet.cell(row, 2).value})
            dictionary.update({'sEmp' + str(compNum) : sheet.cell(row, 4).value})
            dictionary.update({'lat' + str(compNum) : sheet.cell(row, 5).value})
            dictionary.update({'long' + str(compNum) : sheet.cell(row, 6).value})
            dictionary.update({'sales' + str(compNum) : sheet.cell(row, 17).value})
            dictionary.update({'tEmp' + str(compNum) : sheet.cell(row, 20).value})
            dictionary.update({'year' + str(compNum) : sheet.cell(row, 22).value})
            ++compNum
        elif sheet.cell(row, 7).value == str(number):
            dictionary['comp' + str(compNum)] = sheet.cell(row, 3).value
            dictionary.update({'duns' + str(compNum) : sheet.cell(row, 2).value})
            dictionary.update({'sEmp' + str(compNum) : sheet.cell(row, 4).value})
            dictionary.update({'lat' + str(compNum) : sheet.cell(row, 5).value})
            dictionary.update({'long' + str(compNum) : sheet.cell(row, 6).value})
            dictionary.update({'sales' + str(compNum) : sheet.cell(row, 17).value})
            dictionary.update({'tEmp' + str(compNum) : sheet.cell(row, 20).value})
            dictionary.update({'year' + str(compNum) : sheet.cell(row, 22).value})
            ++compNum
    book = load_workbook(filename = "audience1.xlsx")
    sheet = book['Sheet1']
    for row in range(2, 26):
        if (sheet.cell(row, 1)).value == (int)((str(number))[0:2]):
            dictionary.update({'SUPPLY' : sheet.cell(row, 3).value})
            dictionary.update({'DRIVER' : sheet.cell(row, 4).value})
            dictionary.update({'DEMAND' : sheet.cell(row, 5).value})
            dictionary.update({'sec' : sheet.cell(row, 2)})
            break
    #sector data
    dataNum = 0
    year = 7
    sheet = book['Sheet1']
    for i in range (7, 10):
        book = load_workbook(filename = "naicssector_200" + str(i) + "parsed.xlsx")
        for row in range (2, 8989):
            if(dataNum == 10):
                break
            if(sheet.cell(row, 2)) == (str(number))[:2]:
                print("if entered")
                dictionary.update({'FIRMS' + str(dataNum): sheet.cell(row, 4).value})
                dictionary.update({'ESTABLISHMENT' + str(dataNum): sheet.cell(row, 5).value})
                dictionary.update({'EMPLOYMENT' + str(dataNum): sheet.cell(row, 6).value})
                dictionary.update({'ANNUAL PAYROLL' + str(dataNum): sheet.cell(row, 7).value})
                ++dataNum
    for i in range (10, 14):
        book = load_workbook(filename = "naicssector_20" + str(i) + "parsed.xlsx")
        for row in range (2, 8989):
            if(dataNum == 10):
                break

            if(sheet.cell(row, 2)) == (str(number))[:2]:
                dictionary.update({'FIRMS' + str(dataNum): sheet.cell(row, 4).value})
                dictionary.update({'ESTABLISHMENT' + str(dataNum): sheet.cell(row, 5).value})
                dictionary.update({'EMPLOYMENT' + str(dataNum): sheet.cell(row, 6).value})
                dictionary.update({'ANNUAL PAYROLL' + str(dataNum): sheet.cell(row, 7).value})
                ++dataNum
    return dictionary

    
dictionary = routine(541199)
print(dictionary['SUPPLY'])
#drawGraphs(dictionary)